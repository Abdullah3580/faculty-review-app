from openpyxl import load_workbook
import re
import json
import os

# Excel ফাইলের path — আপনার Downloads বা যেখানে আছে সেখানে রাখুন
EXCEL_PATH = r"D:\Desktop\last-file.xlsx"

dept_map = {
    'Internship': 'School of Business and Economics',
    'term 2': 'School of Business and Economics',
    'Fall 2025': 'General',
    'Dept. of Pharmacy': 'Pharmacy',
    'Department of CSE': 'CSE',
    'Department of EEE': 'EEE',
    'Department of Civil Engineering': 'Civil Engineering',
    'School of Business and Economic': 'School of Business and Economics',
    'Department of English': 'English',
    'Department of EDS': 'EDS',
    'Department of MSJ': 'MSJ',
    'Department of BGE': 'BGE',
}

def clean_name(raw_name):
    if not raw_name:
        return None
    cleaned = re.sub(r'^[A-Za-z]{2,6}\s+', '', str(raw_name)).strip()
    return cleaned if cleaned else str(raw_name).strip()

def get_designation(name):
    name_lower = name.lower()
    if 'prof. dr.' in name_lower or 'prof dr' in name_lower:
        return 'Professor'
    elif 'dr.' in name_lower or ' dr ' in name_lower or name_lower.startswith('dr '):
        return 'Assistant Professor'
    else:
        return 'Lecturer'

wb = load_workbook(EXCEL_PATH, read_only=True)
teachers = {}

for sheet_name in wb.sheetnames:
    if sheet_name == 'Summary':
        continue
    ws = wb[sheet_name]
    dept = dept_map.get(sheet_name, 'General')

    for row in ws.iter_rows(values_only=True):
        if not row[0] or not isinstance(row[0], int):
            continue
        raw_name = row[1]
        email = row[2]
        role = row[3]

        if not raw_name or not email:
            continue
        if role and 'non-editing' in str(role).lower():
            continue

        name_str = str(raw_name)
        parts = name_str.split(' ', 1)
        initial = parts[0] if len(parts[0]) <= 6 and not parts[0].lower().startswith(('dr', 'mr', 'ms', 'prof')) else None
        clean = clean_name(raw_name)
        designation = get_designation(clean)
        email_key = str(email).strip().lower()

        if email_key not in teachers:
            teachers[email_key] = {
                'name': clean,
                'email': str(email).strip(),
                'initial': initial,
                'department': dept,
                'designation': designation,
            }

result = list(teachers.values())

output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'faculty_seed.json')
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"✅ Done! {len(result)} faculty saved to {output_path}")